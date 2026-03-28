from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Sheet 1 - Comparaison principale
sheet = wb.active
sheet.title = "Comparaison Serveurs"

# Styles
title_font = Font(name='Times New Roman', size=18, bold=True, color="000000")
header_fill = PatternFill(start_color="1B3F66", end_color="1B3F66", fill_type="solid")
header_font = Font(name='Times New Roman', color="FFFFFF", bold=True, size=11)
cell_font = Font(name='Times New Roman', size=11)
alt_fill = PatternFill(start_color="E9E9E9", end_color="E9E9E9", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Title
sheet['B2'] = "Comparaison des Serveurs de Déploiement Gratuits - Restaurant OS"
sheet['B2'].font = title_font
sheet['B2'].alignment = Alignment(horizontal='left', vertical='center')
sheet.row_dimensions[2].height = 30

# Headers
headers = ["Serveur", "Gratuit", "PostgreSQL", "WebSocket", "HTTPS Auto", "Domaine Custom", "Région Afrique", "Score Global"]
for col, header in enumerate(headers, 2):
    cell = sheet.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Data - Serveurs de déploiement
servers = [
    ["Vercel", "✅ Oui", "✅ (Vercel Postgres)", "⚠️ Limité", "✅ Auto", "✅ Gratuit", "❌ Non", "8.5/10"],
    ["Railway", "✅ $5/mois", "✅ Inclus", "✅ Full", "✅ Auto", "✅ Gratuit", "❌ Non", "8.0/10"],
    ["Render", "✅ Oui", "✅ 90 jours", "✅ Supporté", "✅ Auto", "✅ Gratuit", "❌ Non", "7.5/10"],
    ["Fly.io", "✅ 3 VMs", "✅ Inclus", "✅ Full", "✅ Auto", "✅ Gratuit", "⚠️ Limité", "8.0/10"],
    ["Supabase", "✅ 500MB", "✅ Full", "✅ Temps Réel", "✅ Auto", "✅ Gratuit", "❌ Non", "9.0/10"],
    ["Koyeb", "✅ $5.50/mois", "✅ Inclus", "✅ Full", "✅ Auto", "✅ Gratuit", "❌ Non", "7.5/10"],
    ["Northflank", "✅ Oui", "✅ Inclus", "✅ Full", "✅ Auto", "✅ Gratuit", "❌ Non", "7.0/10"],
    ["Coolify (Self-hosted)", "✅ Illimité", "✅ Inclus", "✅ Full", "✅ Auto", "✅ Gratuit", "✅ Votre serveur", "9.5/10"]
]

for row_idx, server in enumerate(servers, 5):
    for col_idx, value in enumerate(server, 2):
        cell = sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font
        cell.alignment = center_align
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = alt_fill
        # Color coding for status
        if "✅" in str(value):
            cell.fill = green_fill
        elif "❌" in str(value):
            cell.fill = red_fill
        elif "⚠️" in str(value):
            cell.fill = yellow_fill

# Column widths
col_widths = [5, 20, 15, 20, 15, 15, 18, 18, 15]
for i, width in enumerate(col_widths, 1):
    sheet.column_dimensions[get_column_letter(i)].width = width

# Sheet 2 - Détails par serveur
detail_sheet = wb.create_sheet("Détails Complets")

detail_sheet['B2'] = "Détails des Options de Déploiement"
detail_sheet['B2'].font = title_font
detail_sheet['B2'].alignment = Alignment(horizontal='left', vertical='center')
detail_sheet.row_dimensions[2].height = 30

details_headers = ["Serveur", "Avantages", "Inconvénients", "Meilleur Pour", "Prix Production"]
for col, header in enumerate(details_headers, 2):
    cell = detail_sheet.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

details_data = [
    ["Vercel", 
     "Optimisé Next.js\nDéploiement Git auto\nCDN mondial\nSSL automatique\nEdge Functions", 
     "WebSocket limité (serverless)\nPas de datacenter Afrique\nFonctions avec timeout",
     "Frontend Next.js\nAPI REST\nSites statiques",
     "Pro: $20/mois"],
    ["Railway", 
     "Déploiement simple\nPostgreSQL inclus\nWebSocket supporté\nVariables d'env simples\nMonitoring intégré",
     "Crédit limité $5/mois\nPeut devenir cher\nPas de région Afrique",
     "Full-stack apps\nApps temps réel\nMVP rapide",
     "Usage-based\n~$5-20/mois"],
    ["Render", 
     "Gratuit pour web services\nPostgreSQL gratuit\nAuto SSL\nDocker supporté",
     "DB expire 90 jours inactivité\nCold start lent\nLimité ressources gratuites",
     "Démos\nPOC\nPetits projets",
     "Starter: $7/mois"],
    ["Fly.io", 
     "Docker natif\nMultiple régions\nWebSocket full\nPostgreSQL distribué\nCLI puissant",
     "Crédit limité\nConfiguration complexe\nLatence Afrique",
     "Apps distribuées\nMicroservices\nDocker apps",
     "~$5-15/mois"],
    ["Supabase", 
     "PostgreSQL complet\nAuth intégré\nTemps réel natif\nStorage inclus\nAPI auto-générée",
     "Surtout pour backend\nNécessite frontend séparé\n500MB limité",
     "Backend as a Service\nAuth + DB\nTemps réel",
     "Pro: $25/mois"],
    ["Koyeb", 
     "Global deployment\nDocker + Git\nPostgreSQL intégré\nSSL auto\nScaling auto",
     "Crédit limité\nMoins populaire\nDocumentation limitée",
     "Apps globales\nMicroservices",
     "~$5-20/mois"],
    ["Northflank", 
     "CI/CD intégré\nManaged databases\nAuto-scaling\nMulti-cloud",
     "Interface complexe\nCourbe d'apprentissage\nLimites plan gratuit",
     "Teams DevOps\nApps enterprise",
     "Scale: $29/mois"],
    ["Coolify", 
     "100% gratuit\nSelf-hosted\nTous services inclus\nContrôle total\nBackup facile",
     "Nécessite VPS\nMaintenance requise\nPas de support managé",
     "Production complète\nContrôle total\nCoût maîtrisé",
     "Coût VPS\n$5-10/mois"]
]

for row_idx, row_data in enumerate(details_data, 5):
    for col_idx, value in enumerate(row_data, 2):
        cell = detail_sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font
        cell.alignment = left_align
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = alt_fill
    detail_sheet.row_dimensions[row_idx].height = 80

# Column widths for detail sheet
detail_widths = [5, 18, 35, 35, 25, 18]
for i, width in enumerate(detail_widths, 1):
    detail_sheet.column_dimensions[get_column_letter(i)].width = width

# Sheet 3 - Recommandation Restaurant OS
recommend_sheet = wb.create_sheet("Recommandation")

recommend_sheet['B2'] = "Recommandation pour Restaurant OS SaaS"
recommend_sheet['B2'].font = title_font
recommend_sheet['B2'].alignment = Alignment(horizontal='left', vertical='center')
recommend_sheet.row_dimensions[2].height = 30

# Recommendation content
recommendations = [
    ["Option Recommandée", "Architecture", "Coût Estimé", "Complexité"],
    ["🏆 RECOMMANDÉ: Vercel + Supabase", "Frontend Vercel + DB Supabase", "Gratuit → $25/mois", "⭐ Facile"],
    ["🥈 ALTERNATIVE: Railway", "Full-stack sur Railway", "$5-20/mois", "⭐⭐ Moyen"],
    ["🥉 CONTRÔLE TOTAL: Coolify", "Self-hosted sur VPS", "$5-10/mois (VPS)", "⭐⭐⭐ Avancé"]
]

for row_idx, row_data in enumerate(recommendations, 4):
    for col_idx, value in enumerate(row_data, 2):
        cell = recommend_sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font if row_idx > 4 else header_font
        cell.fill = header_fill if row_idx == 4 else (green_fill if row_idx == 5 else alt_fill)
        cell.alignment = center_align
        cell.border = thin_border

# Detailed recommendation
recommend_sheet['B9'] = "Configuration Optimale Recommandée:"
recommend_sheet['B9'].font = Font(name='Times New Roman', size=14, bold=True)

config_details = [
    "",
    "1. FRONTEND (Vercel - Gratuit):",
    "   • Next.js 16 avec optimisations automatiques",
    "   • HTTPS automatique avec certificats SSL",
    "   • CDN mondial pour assets statiques",
    "   • Déploiement continu depuis Git",
    "",
    "2. BASE DE DONNÉES (Supabase - Gratuit 500MB):",
    "   • PostgreSQL managé avec backups",
    "   • Authentification intégrée (optionnelle)",
    "   • Temps réel natif (alternative WebSocket)",
    "   • API REST auto-générée",
    "",
    "3. WEBSOCKET (Option A - Supabase Realtime):",
    "   • Utiliser le temps réel Supabase",
    "   • Changements de DB en temps réel",
    "   • Pas besoin de serveur séparé",
    "",
    "4. WEBSOCKET (Option B - Render/Railway):",
    "   • Déployer le serveur Socket.io séparément",
    "   • Plan gratuit Render ou Railway",
    "",
    "5. STOCKAGE FICHIERS:",
    "   • Supabase Storage (5GB gratuit)",
    "   • Pour images restaurants, menus, etc."
]

for i, line in enumerate(config_details, 10):
    recommend_sheet.cell(row=i, column=2, value=line).font = cell_font

recommend_sheet.column_dimensions['B'].width = 60
recommend_sheet.column_dimensions['C'].width = 30
recommend_sheet.column_dimensions['D'].width = 20
recommend_sheet.column_dimensions['E'].width = 20

# Sheet 4 - Migration PostgreSQL
migration_sheet = wb.create_sheet("Migration PostgreSQL")

migration_sheet['B2'] = "Guide de Migration SQLite → PostgreSQL"
migration_sheet['B2'].font = title_font
migration_sheet['B2'].alignment = Alignment(horizontal='left', vertical='center')
migration_sheet.row_dimensions[2].height = 30

migration_steps = [
    ["Étape", "Action", "Commande/Code", "Notes"],
    ["1", "Installer PostgreSQL", "sudo apt install postgresql", "Ubuntu/Debian"],
    ["2", "Créer la base de données", "CREATE DATABASE restaurant_os;", "Via psql"],
    ["3", "Modifier schema.prisma", 'provider = "postgresql"', "Changer sqlite → postgresql"],
    ["4", "Configurer .env", "DATABASE_URL=\"postgresql://...\"", "URL de connexion"],
    ["5", "Générer client Prisma", "npx prisma generate", "Régénérer le client"],
    ["6", "Créer migrations", "npx prisma migrate dev --name init", "Créer le schéma"],
    ["7", "Migrer données", "npx prisma db seed", "Ou script custom"],
    ["8", "Vérifier intégrité", "npx prisma validate", "Valider le schéma"]
]

for row_idx, row_data in enumerate(migration_steps, 4):
    for col_idx, value in enumerate(row_data, 2):
        cell = migration_sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font if row_idx > 4 else header_font
        cell.fill = header_fill if row_idx == 4 else (alt_fill if row_idx % 2 == 0 else PatternFill())
        cell.alignment = left_align
        cell.border = thin_border

migration_widths = [5, 8, 30, 45, 25]
for i, width in enumerate(migration_widths, 1):
    migration_sheet.column_dimensions[get_column_letter(i)].width = width

# Save workbook
wb.save('/home/z/my-project/download/Comparaison_Serveurs_Deploiement.xlsx')
print("Excel file created successfully!")
