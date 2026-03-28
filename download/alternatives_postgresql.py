from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Sheet 1 - Comparaison
sheet = wb.active
sheet.title = "Alternatives PostgreSQL"

# Styles
title_font = Font(name='Times New Roman', size=18, bold=True, color="000000")
header_fill = PatternFill(start_color="1B3F66", end_color="1B3F66", fill_type="solid")
header_font = Font(name='Times New Roman', color="FFFFFF", bold=True, size=11)
cell_font = Font(name='Times New Roman', size=11)
alt_fill = PatternFill(start_color="E9E9E9", end_color="E9E9E9", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Title
sheet['B2'] = "Alternatives Gratuites à Supabase pour PostgreSQL"
sheet['B2'].font = title_font
sheet['B2'].alignment = Alignment(horizontal='left', vertical='center')
sheet.row_dimensions[2].height = 30

# Headers
headers = ["Service", "Stockage", "Limite", "WebSocket", "Temps Réel", "API Auto", "Régions", "Score"]
for col, header in enumerate(headers, 2):
    cell = sheet.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Data - PostgreSQL alternatives
services = [
    ["🏆 NEON", "0.5 GB", "Illimité", "❌ Non", "❌ Non", "❌ Non", "✅ USA/EU", "9.5/10"],
    ["Supabase", "500 MB", "Illimité", "✅ Oui", "✅ Oui", "✅ Oui", "✅ Global", "9.0/10"],
    ["Render PostgreSQL", "90 jours", "Limité", "❌ Non", "❌ Non", "❌ Non", "✅ USA/EU", "6.5/10"],
    ["Railway", "$5/mois", "Crédit", "✅ Oui", "❌ Non", "❌ Non", "✅ USA/EU", "7.0/10"],
    ["CockroachDB", "10 GB", "Illimité", "❌ Non", "❌ Non", "✅ Oui", "✅ Global", "8.0/10"],
    ["ElephantSQL", "20 MB", "5 connexions", "❌ Non", "❌ Non", "❌ Non", "✅ Global", "5.0/10"],
    ["Fly.io Postgres", "3 VMs", "Limité", "✅ Oui", "❌ Non", "❌ Non", "⚠️ Limité", "7.5/10"],
    ["Aiven", "1 mois essai", "Limité", "❌ Non", "❌ Non", "❌ Non", "✅ Global", "6.0/10"],
]

for row_idx, service in enumerate(services, 5):
    for col_idx, value in enumerate(service, 2):
        cell = sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font
        cell.alignment = center_align
        cell.border = thin_border
        if "🏆" in str(value):
            cell.fill = green_fill
        elif "✅" in str(value):
            cell.fill = green_fill
        elif "❌" in str(value):
            cell.fill = red_fill
        elif "⚠️" in str(value):
            cell.fill = yellow_fill
        elif row_idx % 2 == 0:
            cell.fill = alt_fill

# Column widths
widths = [5, 22, 12, 12, 12, 12, 12, 14, 12]
for i, width in enumerate(widths, 1):
    sheet.column_dimensions[get_column_letter(i)].width = width

# Sheet 2 - Détails NEON
neon_sheet = wb.create_sheet("NEON - Recommandé")

neon_sheet['B2'] = "NEON - Meilleure Alternative Gratuite PostgreSQL"
neon_sheet['B2'].font = title_font
neon_sheet['B2'].alignment = Alignment(horizontal='left', vertical='center')
neon_sheet.row_dimensions[2].height = 30

neon_details = [
    ["Caractéristique", "Détail"],
    ["Stockage gratuit", "0.5 GB (extensible)"],
    ["Compute", "191.9 hours/mois gratuit"],
    ["Projets", "1 projet gratuit"],
    ["Branches", "10 branches par projet"],
    ["Autoscaling", "✅ Oui"],
    ["Auto-suspend", "✅ Après 5 min inactivité"],
    ["Connection pooling", "✅ Inclus"],
    ["SSL", "✅ Automatique"],
    ["Backups", "✅ 7 jours retention"],
    ["Régions", "AWS us-east-2, eu-central-1, ap-southeast-1"],
    ["API REST", "❌ Non (à configurer soi-même)"],
    ["WebSocket", "❌ Non inclus"],
]

for row_idx, row_data in enumerate(neon_details, 4):
    for col_idx, value in enumerate(row_data, 2):
        cell = neon_sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font if row_idx > 4 else header_font
        cell.fill = header_fill if row_idx == 4 else (alt_fill if row_idx % 2 == 0 else PatternFill())
        cell.alignment = left_align
        cell.border = thin_border

neon_widths = [5, 25, 50]
for i, width in enumerate(neon_widths, 1):
    neon_sheet.column_dimensions[get_column_letter(i)].width = width

# Sheet 3 - Architecture Recommandée
arch_sheet = wb.create_sheet("Architecture Recommandée")

arch_sheet['B2'] = "Architecture Optimale Gratuite - Restaurant OS"
arch_sheet['B2'].font = title_font
arch_sheet['B2'].alignment = Alignment(horizontal='left', vertical='center')
arch_sheet.row_dimensions[2].height = 30

arch_data = [
    ["Composant", "Service", "Fonction", "Coût"],
    ["Frontend", "Vercel", "Next.js hosting + CDN + SSL", "GRATUIT"],
    ["Base de données", "NEON", "PostgreSQL serverless", "GRATUIT"],
    ["WebSocket/Realtime", "Pusher (ou Ably)", "Temps réel pour commandes", "GRATUIT (200k msgs/j)"],
    ["Stockage fichiers", "Cloudinary", "Images restaurants/menus", "GRATUIT (25GB)"],
    ["Emails", "Resend", "Notifications email", "GRATUIT (3000/mois)"],
    ["Authentification", "Custom + NextAuth", "Auth intégrée au projet", "GRATUIT"],
]

for row_idx, row_data in enumerate(arch_data, 4):
    for col_idx, value in enumerate(row_data, 2):
        cell = arch_sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font if row_idx > 4 else header_font
        cell.fill = header_fill if row_idx == 4 else (green_fill if "GRATUIT" in str(value) else alt_fill)
        cell.alignment = center_align
        cell.border = thin_border

arch_widths = [5, 18, 20, 35, 20]
for i, width in enumerate(arch_widths, 1):
    arch_sheet.column_dimensions[get_column_letter(i)].width = width

# Sheet 4 - Comparaison Détaillée
compare_sheet = wb.create_sheet("Pourquoi NEON")

compare_sheet['B2'] = "Pourquoi NEON est le meilleur choix gratuit"
compare_sheet['B2'].font = title_font
compare_sheet['B2'].alignment = Alignment(horizontal='left', vertical='center')
compare_sheet.row_dimensions[2].height = 30

compare_content = [
    ["Avantage NEON", "Explication"],
    ["Serverless", "Pas de serveur à gérer, scale automatique"],
    ["Auto-suspend", "Économise les ressources quand inactif"],
    ["Branches", "Créer des branches DB pour les tests (comme Git)"],
    ["Connection Pooling", "Inclus nativement, idéal pour serverless"],
    ["PostgreSQL pur", "100% compatible PostgreSQL standard"],
    ["Pas de limite de temps", "Contrairement à Render (90 jours)"],
    ["Pas de crédit limité", "Contrairement à Railway ($5/mois)"],
    ["Régions EU", "Conformité RGPD pour clients africains"],
    ["CLI puissant", "neon CLI pour gérer les projets"],
    ["Intégration Vercel", "Intégration native avec Vercel"],
]

for row_idx, row_data in enumerate(compare_content, 4):
    for col_idx, value in enumerate(row_data, 2):
        cell = compare_sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font if row_idx > 4 else header_font
        cell.fill = header_fill if row_idx == 4 else (alt_fill if row_idx % 2 == 0 else PatternFill())
        cell.alignment = left_align
        cell.border = thin_border

compare_widths = [5, 25, 55]
for i, width in enumerate(compare_widths, 1):
    compare_sheet.column_dimensions[get_column_letter(i)].width = width

# Save
wb.save('/home/z/my-project/download/Alternatives_PostgreSQL_Gratuites.xlsx')
print("Excel file created successfully!")
